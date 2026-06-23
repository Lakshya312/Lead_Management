from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import (Product, Product_Category, Region, Territory,
                     Lead, Lead_Source, Lead_Status, UserProfile)
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib import messages
from django.http import JsonResponse
import json
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.db.models.deletion import ProtectedError
from django.utils import timezone
from functools import wraps
from openpyxl import load_workbook

# ================ DECORATOR ================
def allowed_roles(roles=[]):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):

            if not request.user.is_authenticated:
                return redirect("login")

            user_profile, created = UserProfile.objects.get_or_create(
                user=request.user,
                defaults={"role": "Executive"}
            )

            user_role = user_profile.role

            if user_role in roles:
                return view_func(request, *args, **kwargs)

            messages.error(request, "You are not authorized to access this page.")
            return redirect("dashboard")

        return wrapper
    return decorator


# ================ HOME ================
def home(request):
    context = {
        "lead_count": Lead.objects.count(),
        "product_count": Product.objects.count(),
        "region_count": Region.objects.count(),
        "source_count": Lead_Source.objects.count(),
    }
    return render(request, "home.html", context)

# ================= DASHBOARD =================
@login_required
def dashboard(request):
    try:
        context = {
            'product_count': Product.objects.count(),
            'region_count': Region.objects.count(),
            "lead_count":Lead.objects.count(),
                #filter(Added_By=request.user.username)
        }
        return render(request, 'dashboard.html', context)

    except Exception as e:
        messages.error(request, "Dashboard could not be load")
        return redirect("home")


# ================= PRODUCT =================
@allowed_roles(["Admin", "Manager", "Executive"])
def product_list(request):
    search = request.GET.get('search')

    products = Product.objects.all()

    if search:
        products = products.filter(
            Q(ProductName__icontains=search)
        )

    return render(
        request,
        'product/product_list.html',
        {'products': products}
    )

@allowed_roles(["Admin", "Manager"])
def add_product(request):

    categories = Product_Category.objects.all()

    if request.method == "POST":

        product_name = request.POST.get("ProductName")
        category_id = request.POST.get("CategoryID")
        is_active = request.POST.get("Is_Active")

        # validation
        if not product_name:
            messages.error(request, "Product Name cannot be empty")
            return redirect("add_product")

        if not category_id:
            messages.error(request, "Category must be selected")
            return redirect("add_product")

        try:
            category = Product_Category.objects.get(CategoryID=category_id)
        except Product_Category.DoesNotExist:
            messages.error(request, "Invalid Category selected")
            return redirect("add_product")

        # Convert True/False string to Boolean
        is_active = True if is_active == "True" else False

        # save product
        try:
            Product.objects.create(
                ProductName=product_name,
                Category=category,
                Is_Active=is_active,
                Added_By=request.user.username
            )

            messages.success(request, "Product added successfully")
            return redirect("product_list")

        except Exception as e:
            messages.error(request, "Product could not be load")
            return redirect("add_product")

    return render(request, "product/add_product.html", {
        "categories": categories
    })

@allowed_roles(["Admin", "Manager"])
def upload_products_excel(request):

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("product_list")

        if not excel_file.name.endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are allowed.")
            return redirect("product_list")

        wb = load_workbook(excel_file)
        sheet = wb.active

        errors = []

        for row_no, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True),
                start=2):

            product_name = row[0]
            category_name = row[1]
            is_active_value = row[2]

            # Product Name Validation
            if not product_name or str(product_name).strip() == "":
                errors.append(
                    f"Row {row_no}: Product Name cannot be empty."
                )
                continue

            # Category Validation
            if not category_name or str(category_name).strip() == "":
                errors.append(
                    f"Row {row_no}: Category cannot be empty."
                )
                continue

            category, created = Product_Category.objects.get_or_create(
                CategoryName=str(category_name).strip(),
                defaults={
                    "Added_By": request.user.username
                }
            )

            # IsActive Validation
            if is_active_value is None or str(is_active_value).strip() == "":
                is_active = True
            else:
                is_active = str(
                    is_active_value
                ).strip().lower() in [
                                "true", "yes", "1", "active"
                            ]

            # Duplicate Validation
            if Product.objects.filter(
                    ProductName=str(product_name).strip(),
                    Category=category
            ).exists():
                errors.append(
                    f"Row {row_no}: Product '{product_name}' already exists."
                )
                continue

            Product.objects.create(
                ProductName=str(product_name).strip(),
                Category=category,
                Is_Active=is_active,
                Added_By=request.user.username
            )

    if errors:
        for error in errors:
            messages.error(request, error)
    else:
        messages.success(
            request,
            "Products uploaded successfully."
        )

        #messages.success(request, "Products uploaded successfully.")
        #return redirect("product_list")

    return redirect("product_list")

@allowed_roles(["Admin", "Manager"])
def edit_product(request, id):

    if request.user.userprofile.role == "Admin":
        product = Product.objects.filter(ProductID=id).first()
    else:
        product = Product.objects.filter(
            ProductID=id,
            Added_By=request.user.username
        ).first()

    if not product:
        messages.error(request, "Permission denied. You are not allowed to edit this product.")
        return redirect("product_list")

    categories = Product_Category.objects.all()

    if request.method == "POST":
        product.ProductName = request.POST.get("ProductName")

        category_id = request.POST.get("CategoryID")
        category = get_object_or_404(Product_Category, CategoryID=category_id)

        product.Category = category
        product.Is_Active = request.POST.get("Is_Active") == "True"

        product.save()

        messages.success(request, "Product updated successfully.")
        return redirect("product_list")

    return render(request, "product/edit_product.html", {
        "product": product,
        "categories": categories
    })

@allowed_roles(["Admin", "Manager"])
def delete_product(request, id):
    try:
        product = get_object_or_404(Product, ProductID=id, Added_By=request.user.username)

        try:
            product.delete()
            messages.success(request, "Product deleted successfully.")

        except ProtectedError:
            messages.error(
                request,
                "Product cannot be deleted because it is used in one or more leads."
            )

    except Product.DoesNotExist:
        messages.error(request, "Product not found.")

    return redirect("product_list")

@csrf_exempt
def product_api(request):

    if request.method == "GET":
        products = list(
            Product.objects.values(
                'ProductID',
                'ProductName',
                'Category_id',
                'Category__CategoryName',
            )
        )
        return JsonResponse({"success": True, "data": products})


    elif request.method == "POST":
        try:
            data = json.loads(request.body)

            product_name = data.get("ProductName")
            category_id = data.get("CategoryID")

            if not product_name or not category_id:
                return JsonResponse({
                    "success": False,
                    "message": "ProductName and CategoryID are required"
                }, status=400)

            category = Product_Category.objects.get(CategoryID=category_id)

            product = Product.objects.create(
                ProductName=product_name,
                Category=category
            )

            return JsonResponse({
                "success": True,
                "message": "Product created successfully",
                "data": {
                    "ProductID": product.ProductID,
                    "ProductName": product.ProductName,
                    "CategoryID": product.Category.CategoryName
                }
            })

        except Product_Category.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": "Invalid CategoryID"
            }, status=400)

        except json.JSONDecodeError:
            return JsonResponse({
                "success": False,
                "message": "Invalid JSON format"
            }, status=400)

@csrf_exempt
def product_detail_api(request, id):

    if request.method == "GET":
        try:
            product = Product.objects.get(ProductID=id)

            data = {
                "ProductID": product.ProductID,
                "ProductName": product.ProductName,
                "Category": product.Category.CategoryName if product.Category else None,
                "Is_Active": product.Is_Active,
            }

            return JsonResponse({"success": True, "data": data})

        except Product.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": "Product not found"
            },status=404)


    elif request.method == "DELETE":
        try:
            product = Product.objects.get(ProductID=id)
            try:
                product.delete()
                return JsonResponse({
                    "success": True,
                    "message": "Product deleted successfully."
                }, status=200)

            except ProtectedError:
                return JsonResponse({
                "success": False,
                "message": "Product cannot be deleted because it is used in one or more leads."
            }, status=400)

        except Product.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": "Product not found."
            }, status=404)


# ================= REGION =================
@allowed_roles(["Admin", "Manager", "Executive"])
def region_list(request):
    search = request.GET.get('search')

    regions = Region.objects.all()

    if search:
        regions = regions.filter(
            RegionName__icontains=search
        )

    return render(
        request,
        'region/region_list.html',
        {'regions': regions}
    )

@allowed_roles(["Admin", "Manager"])
def add_region(request):
    if request.method == "POST":
        try:
            region_name = request.POST.get("RegionName")

            if not region_name:
                messages.error(request, "Region Name cannot be empty")
                return redirect("add_region")

            Region.objects.create(RegionName=region_name)
            messages.success(request, "Region added successfully")
            return redirect("region_list")

        except Exception as e:
            messages.error(request, "The region could not be added.")
            return redirect("add_region")

    return render(request, "region/add_region.html", {
        "regions": Region.objects.all()
    })

@allowed_roles(["Admin", "Manager"])
def edit_region(request, id):
    region = get_object_or_404(Region, RegionID=id, Added_By=request.user.username)

    if request.method == "POST":
        region.RegionName = request.POST.get("RegionName")
        region.save()
        return redirect("region_list")

    return render(request, "region/edit_region.html", {"region": region})

@allowed_roles(["Admin", "Manager"])
def delete_region(request, id):
    region = get_object_or_404(Region, RegionID=id, Added_By=request.user.username)
    region.delete()
    return redirect("region_list")


@csrf_exempt
def region_api(request):

    if request.method == "GET":
        regions = list(Region.objects.values())

        return JsonResponse({
            "success": True,
            "data": regions
        })

    elif request.method == "POST":
        try:
            data = json.loads(request.body)

            region_name = data.get("RegionName")

            if not region_name:
                return JsonResponse({
                    "success": False,
                    "message": "RegionName is required"
                }, status=400)

            region = Region.objects.create(
                RegionName=region_name
            )

            return JsonResponse({
                "success": True,
                "message": "Region created successfully",
                "data": {
                    "RegionID": region.RegionID,
                    "RegionName": region.RegionName
                }
            }, status=201)

        except json.JSONDecodeError:
            return JsonResponse({
                "success": False,
                "message": "Invalid JSON"
            }, status=400)

@csrf_exempt
def region_detail_api(request, id):

    if request.method == "GET":
        try:
            region = Region.objects.get(RegionID=id)

            data = {
                "RegionID": region.RegionID,
                "RegionName": region.RegionName,
            }

            return JsonResponse({
                "success": True,
                "data": data
            })

        except Region.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": "Region not found"
            }, status=404)

    elif request.method == "DELETE":
        try:
            region = Region.objects.get(RegionID=id)
            region.delete()

            return JsonResponse({
                "success": True,
                "message": "Region deleted successfully"
            })

        except Region.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": "Region not found"
            }, status=404)


# ================= LEAD =================
@allowed_roles(["Admin", "Manager", "Executive"])
def lead_list(request):
    search = request.GET.get('search')

    role = request.user.userprofile.role

    if role == "Executive":
        leads = Lead.objects.filter(
            Added_By=request.user.username
        )
    else:
        leads = Lead.objects.all()

    leads = leads.select_related(
        'Region',
        'Source',
        'Product',
        'Status',
        'Territory'
    ).order_by('LeadID')

    if search:
        leads = leads.filter(
            Q(PersonName__icontains=search) |
            Q(ContactNo__icontains=search) |
            Q(Email__icontains=search) |
            Q(CompanyName__icontains=search)
        )

    paginator = Paginator(leads, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "new_count": leads.filter(Status_id=1).count(),
        "qualified_count": leads.filter(Status_id=2).count(),
        "lost_count": leads.filter(Status_id=3).count(),
        "search": search,
    }

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return render(request, "lead/lead_table.html", context)

    return render(request, "lead/lead_list.html", context)


@allowed_roles(["Admin", "Manager", "Executive"])
def add_lead(request):
    if request.method == "POST":

        person_name = request.POST.get("PersonName")
        Gender = request.POST.get("Gender")

        if not person_name or not person_name.strip():
            messages.error(request, "Person Name cannot be empty")
            return redirect("add_lead")

        if not Gender:
            messages.error(request, "Gender cannot be empty")
            return redirect("add_lead")

        email = request.POST.get("Email")

        if not email or not email.endswith(".com"):
            messages.error(request, "Please enter a valid .com email address")
            return redirect("add_lead")

        try:
            Lead.objects.create(
                PersonName=person_name.strip(),
                Gender=Gender,
                Email=request.POST.get("Email"),
                ContactNo=request.POST.get("ContactNo"),
                CompanyName=request.POST.get("CompanyName"),
                BusinessNeed=request.POST.get("BusinessNeed") or "Not Specified",
                City=request.POST.get("City"),
                State=request.POST.get("State"),

                Product_id=request.POST.get("Product"),
                Region_id=request.POST.get("Region"),
                Territory_id=request.POST.get("Territory"),
                Source_id=request.POST.get("Source"),
                Status_id=request.POST.get("Status"),
                Lead_Gen_Date=request.POST.get("Lead_Gen_Date"),
                ExecutiveID=request.user.id,
                Added_By=request.user.username,
            )

            messages.success(request, "Lead added successfully")
            return redirect("lead_list")

        except Exception as e:
            messages.error(request, "The lead could not be added.")
            return redirect("add_lead")

    return render(request, "lead/add_lead.html", {
        "products": Product.objects.all(),
        "regions": Region.objects.all(),
        "territories": Territory.objects.all(),
        "sources": Lead_Source.objects.all(),
        "statuses": Lead_Status.objects.all(),
        "genders": Lead.objects.all(),
    })

@allowed_roles(["Admin", "Manager", "Executive"])
def edit_lead(request, id):

    role = request.user.userprofile.role

    if role == "Executive":
        lead = get_object_or_404(Lead,LeadID=id,Added_By=request.user.username)

    else:
         lead = get_object_or_404(Lead,LeadID=id)

    if request.method == "POST":
        lead.PersonName = request.POST.get("PersonName")
        lead.Email = request.POST.get("Email")
        lead.ContactNo = request.POST.get("ContactNo")
        lead.CompanyName = request.POST.get("CompanyName")
        lead.City = request.POST.get("City")
        lead.State = request.POST.get("State")

        lead.Product_id = request.POST.get("Product")
        lead.Region_id = request.POST.get("Region")
        lead.Source_id = request.POST.get("Source")
        lead.Status_id = request.POST.get("Status")

        lead.save()

        messages.success(request, "Lead updated successfully.")
        return redirect("lead_list")

    return render(request, "lead/edit_lead.html", {
        "lead": lead,
        "products": Product.objects.all(),
        "regions": Region.objects.all(),
        "sources": Lead_Source.objects.all(),
        "statuses": Lead_Status.objects.all(),
    })

@allowed_roles(["Admin", "Manager", "Executive"])
def delete_lead(request, id):
    lead = get_object_or_404(Lead, LeadID=id, Added_By=request.user.username)
    lead.delete()
    return redirect("lead_list")


@csrf_exempt
def lead_api(request):

    if request.method == "GET":
        leads = list(Lead.objects.values())
        return JsonResponse({
            "success": True,
            "data": leads
        })

    elif request.method == "POST":
        try:
            data = json.loads(request.body)

            person_name = data.get("PersonName")

            if not person_name:
                return JsonResponse({
                    "success": False,
                    "message": "PersonName is required"
                }, status=400)

            lead = Lead.objects.create(
                PersonName=person_name,
                Gender=data.get("Gender"),
                CompanyName=data.get("CompanyName"),
                ContactNo=data.get("ContactNo"),
                Email=data.get("Email"),
                City=data.get("City"),
                State=data.get("State"),
                BusinessNeed=data.get("BusinessNeed"),
                ExecutiveID=data.get("ExecutiveID"),

                Region_id=data.get("RegionID"),
                Product_id=data.get("ProductID"),
                Status_id=data.get("StatusID"),
                Source_id=data.get("LeadSourceID"),
                Territory_id=data.get("TerritoryID")
            )

            return JsonResponse({
                "success": True,
                "message": "Lead created successfully",
                "LeadID": lead.LeadID
            }, status=201)

        except json.JSONDecodeError:
            return JsonResponse({
                "success": False,
                "message": "Invalid JSON"
            }, status=400)

        except Exception as e:
            return JsonResponse({
                "success": False,
                "message": str(e)
            }, status=500)

@csrf_exempt
def lead_detail_api(request, id):

    if request.method == "GET":
        try:
            lead = Lead.objects.get(LeadID=id)

            data = {
                "LeadID": lead.LeadID,
                "PersonName": lead.PersonName,
                "Email": lead.Email,
                "ContactNo": lead.ContactNo,
                "CompanyName": lead.CompanyName,
                "City": lead.City,
                "State": lead.State,
                "BusinessNeed": lead.BusinessNeed,
                "Status": lead.Status.StatusName if lead.Status else None,
                "Product": lead.Product.ProductName if lead.Product else None,
                "Region": lead.Region.RegionName if lead.Region else None,
            }

            return JsonResponse({
                "success": True,
                "data": data
            })

        except Lead.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": "Lead not found"
            }, status=404)

    elif request.method == "DELETE":
        try:
            lead = Lead.objects.get(LeadID=id)
            lead.delete()

            return JsonResponse({
                "success": True,
                "message": "Lead deleted successfully"
            })

        except Lead.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": "Lead not found"
            }, status=404)


# ================= LEAD SOURCE =================
def source_list(request):
    search = request.GET.get('search')

    sources = Lead_Source.objects.all()

    if search:
        sources = sources.filter(
            LeadSourceName__icontains=search
        )

    return render(
        request,
        'source/source_list.html',
        {'sources': sources}
    )


def add_source(request):
    if request.method == "POST":

        if not request.POST.get('LeadSourceName'):
            messages.error(request, "Lead Source Name cannot be empty")
            return redirect('add_source')

        Lead_Source.objects.create(
            LeadSourceName=request.POST.get("LeadSourceName")
        )
        return redirect("source_list")

    return render(request, "source/add_source.html")


def edit_source(request, id):
    source = get_object_or_404(Lead_Source, LeadSourceid=id)

    if request.method == "POST":
        source.LeadSourceName = request.POST.get("LeadSourceName")
        source.save()
        return redirect("source_list")

    return render(request, "source/edit_source.html", {"source": source})


def delete_source(request, id):
    source = get_object_or_404(Lead_Source, LeadSourceid=id)
    source.delete()
    return redirect("source_list")


# ================ REGISTER ================
import re

def register(request):

    if request.method == "POST":

        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")

        # Password Validation
        password_errors = []

        if len(password) < 8:
            password_errors.append(
                "Password must contain at least 8 characters."
            )

        if not re.search(r'[A-Z]', password):
            password_errors.append(
                "One capital letter is required."
            )

        if not re.search(r'[a-z]', password):
            password_errors.append(
                "One small letter is required."
            )

        if not re.search(r'[0-9]', password):
            password_errors.append(
                "One number is required."
            )

        if not re.search(r'[@$!%*?&#]', password):
            password_errors.append(
                "One special character is required."
            )

        if password_errors:
            messages.error(
                request,
                "<ul><li>" + "</li><li>".join(password_errors) + "</li></ul>"
            )
            return redirect("register")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists")
            return redirect("register")

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists")
            return redirect("register")

        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password
            )

            UserProfile.objects.create(
                user=user,
                role="Executive"
            )

            messages.success(request, "Account created successfully")
            return redirect("login")

        except Exception as e:
            messages.error(request, "The account could not be created.")
            return redirect("register")

    return render(request, "register.html")


# ================ LOGIN ================
def user_login(request):

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        if not User.objects.filter(username=username).exists():
            return render(request, "login.html", {
                "username": username,
                "username_error": "Invalid username"
            })

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            # Force current login time update
            user.last_login = timezone.now()
            user.save(update_fields=["last_login"])

            if request.POST.get("remember_me"):
                request.session.set_expiry(60 * 60 * 24 * 7)
            else:
                request.session.set_expiry(0)

            return redirect("dashboard")

        return render(request, "login.html", {
            "username": username,
            "password_error": "Invalid password"
        })

    return render(request, "login.html")

# ================ LOGOUT ===================
def user_logout(request):
    logout(request)
    return redirect("home")

# ================ PROFILE ================
def profile(request):
    user_profile, created = UserProfile.objects.get_or_create(
        user=request.user
    )

    return render(request, "profile.html",
                  {"user_profile": user_profile}
                 )


# ================ EDIT PROFILE ================
def edit_profile(request):
    if request.method == "POST":
        request.user.username = request.POST.get("username")
        request.user.email = request.POST.get("email")
        request.user.save()

        return redirect("profile")

    return render(request, "edit_profile.html")


# ================ USER MANAGEMENT ================
@allowed_roles(["Admin"])
def user_management(request):
    users = User.objects.all().order_by("id")

    for user in users:
        UserProfile.objects.get_or_create(
            user=user,
            defaults={"role": "Executive"}
        )

    return render(request, "users/user_management.html", {
        "users": users
    })


# ================= ROLE UPDATE ================
@allowed_roles(["Admin"])
def update_user_role(request, user_id):

    if request.method == "POST":
        role = request.POST.get("role")

        user = get_object_or_404(User, id=user_id)

        profile, created = UserProfile.objects.get_or_create(
            user=user,
            defaults={"role": "Executive"}
        )

        if profile.is_main_admin:
            messages.error(
                request,
                "Main Admin role cannot be changed."
            )
            return redirect("user_management")

        if role in ["Admin", "Manager", "Executive"]:
            profile.role = role
            profile.save()
            messages.success(request, "User role updated successfully.")
        else:
            messages.error(request, "Invalid role selected.")

    return redirect("user_management")


# ================ USER DELETE ================
@allowed_roles(["Admin"])
def delete_user(request, user_id):

    if request.method == "POST":

        user = get_object_or_404(User, id=user_id)

        if user == request.user:
            messages.error(
                request,
                "You cannot delete your own account."
            )
            return redirect("user_management")

        profile, created = UserProfile.objects.get_or_create(
            user=user,
            defaults={"role": "Executive"}
        )

        if profile.is_main_admin:
            messages.error(
                request,
                "Main Admin account cannot be deleted."
            )
            return redirect("user_management")

        if profile.role == "Admin":

            admin_count = UserProfile.objects.filter(
                role="Admin"
            ).count()

            if admin_count <= 1:
                messages.error(
                    request,
                    "The last Admin account cannot be deleted."
                )
                return redirect("user_management")

        user.delete()

        messages.success(
            request,
            "User deleted successfully."
        )

    return redirect("user_management")